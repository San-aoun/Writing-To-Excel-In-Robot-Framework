from openpyxl import Workbook, load_workbook
from contextlib import closing

def make_excel_file(file_name):
    with closing(Workbook()) as wb:
        wb.save(file_name)

def write_to_excel(file_name, sheet_name, cell_cords, value):
    with closing(load_workbook(filename=file_name)) as wb:
        ws = wb.active
        ws = wb.get_sheet_by_name(name = sheet_name) 
        ws[cell_cords] = value
        wb.save(file_name)
